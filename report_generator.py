import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def main():
    # Read data from CSV
    df = pd.read_csv('sales_data.csv')

    # Write raw data, summary stats, and pivot table to Excel
    with pd.ExcelWriter('report.xlsx', engine='openpyxl') as writer:
        # Raw data sheet
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Summary Statistics sheet
        summary_df = pd.DataFrame({
            'Metric': ['Count', 'Total Sales', 'Average Sale', 'Minimum Sale', 'Maximum Sale'],
            'Value': [
                df['Sales'].count(),
                df['Sales'].sum(),
                df['Sales'].mean(),
                df['Sales'].min(),
                df['Sales'].max()
            ]
        })
        summary_df.to_excel(writer, sheet_name='Summary Stats', index=False)

        # Pivot Table sheet (Total Sales by Category)
        pivot_df = df.groupby('Category')['Sales'].sum().reset_index()
        pivot_df.columns = ['Category', 'Total Sales']
        pivot_df.to_excel(writer, sheet_name='Pivot Table', index=False)

    # Load the workbook to add the chart
    wb = load_workbook('report.xlsx')
    ws = wb['Pivot Table']

    # Create a bar chart for total sales by category
    chart = BarChart()
    chart.title = "Total Sales by Category"
    chart.y_axis.title = 'Total Sales'
    chart.x_axis.title = 'Category'

    # Data range for the chart (including header for series title)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 10

    # Place the chart on the Pivot Table sheet
    ws.add_chart(chart, "E2")

    # Save the final Excel report
    wb.save('report.xlsx')

if __name__ == "__main__":
    main()
